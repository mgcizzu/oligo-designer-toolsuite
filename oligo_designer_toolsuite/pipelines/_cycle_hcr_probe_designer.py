############################################
# imports
############################################

import csv
import logging
import os
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from argparse import ArgumentParser, RawDescriptionHelpFormatter

import pandas as pd
import yaml
from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqUtils import MeltingTemp as mt
from openpyxl import load_workbook

try:
    from seqfold import dg as seqfold_dg
except ImportError:  # pragma: no cover - depends on environment
    seqfold_dg = None


############################################
# data structures
############################################


@dataclass
class CandidateWindow:
    gene: str
    transcript_id: str
    window_start: int
    window_end: int
    target_left_45: str
    target_gap_2: str
    target_right_45: str
    gc_left: float
    gc_right: float
    tm_left: float
    tm_right: float
    rna_tm_left: float
    rna_tm_right: float
    dg_left: float | None
    dg_right: float | None
    max_cross_hybridization_tm_left: float | None
    max_cross_hybridization_tm_right: float | None
    junction_sequence: str
    score: float


############################################
# CycleHCR Probe Designer
############################################


class CycleHCRProbeDesigner:
    """
    Design CycleHCR primary probes from transcript windows and explicit barcode assignments.

    The first working implementation focuses on:
    - transcript sliding-window target selection with deterministic scoring
    - workbook-backed barcode resolution
    - direct and Twist/T7/RT output assembly
    - exact transcriptome uniqueness filtering for the configured junction sequence

    It intentionally avoids optional heavyweight alignment/AI-specific code paths.
    """

    def __init__(self, write_intermediate_steps: bool, dir_output: str, n_jobs: int) -> None:
        self.dir_output = os.path.abspath(dir_output)
        Path(self.dir_output).mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now()
        file_logger = os.path.join(
            self.dir_output,
            f"log_cycle_hcr_probe_designer_{timestamp.year}-{timestamp.month}-{timestamp.day}-{timestamp.hour}-{timestamp.minute}.txt",
        )
        logging.getLogger("log_name")
        logging.basicConfig(
            format="%(asctime)s [%(levelname)s] %(message)s",
            level=logging.NOTSET,
            handlers=[logging.FileHandler(file_logger)],
            force=True,
        )
        logging.captureWarnings(True)
        logging.info("--------------START PIPELINE--------------")

        self.write_intermediate_steps = write_intermediate_steps
        self.n_jobs = n_jobs

    ############################################
    # high-level API
    ############################################

    def run(self, config: dict, gene_ids: list[str] | None) -> dict[str, str]:
        if seqfold_dg is None:
            warnings.warn(
                "seqfold is not available. Secondary-structure filtering is skipped in this environment.",
                UserWarning,
            )

        transcript_records = self._load_transcript_records(
            files_fasta=config["files_fasta_target_probe_database"], gene_ids=gene_ids
        )
        reference_records = self._load_reference_sequences(
            files_fasta=config["files_fasta_reference_database_target_probe"]
        )
        barcode_library = self._load_barcode_library(config["barcode_library"])
        assignments = self._load_assignments(config["barcode_assignment"], barcode_library, gene_ids)
        junction_counts = self._build_junction_counts(reference_records, config["cycle_hcr"])
        cross_hybridization_index = None
        if self._cross_hybridization_tm_screen_enabled(config):
            cross_hybridization_index = self._build_cross_hybridization_index(
                reference_records=reference_records,
                oligo_length=config["cycle_hcr"]["left_probe_length"],
                seed_length=config["cycle_hcr"].get("cross_hybridization_seed_length", 12),
            )
            logging.info("Optional cross-hybridization Tm screen is enabled.")
        else:
            logging.info("Optional cross-hybridization Tm screen is disabled.")

        selected_candidates = []
        rejected_rows = []
        assignment_rows = []

        for gene, transcript_record in transcript_records.items():
            if gene not in assignments:
                rejected_rows.append(
                    {
                        "Gene": gene,
                        "Transcript_ID": transcript_record["transcript_id"],
                        "Window_Start": "",
                        "Window_End": "",
                        "Reason": "missing_barcode_assignment",
                    }
                )
                continue

            assignment = assignments[gene]
            assignment_rows.append(
                {
                    "Gene": gene,
                    "Sheet": assignment["Sheet"],
                    "Left_Barcode_Name": assignment["Left_Barcode_Name"],
                    "Right_Barcode_Name": assignment["Right_Barcode_Name"],
                    "Left_Barcode_Sequence": assignment["left_sequence"],
                    "Right_Barcode_Sequence": assignment["right_sequence"],
                }
            )

            candidates_gene, rejected_gene = self._enumerate_candidates_for_gene(
                gene=gene,
                transcript_id=transcript_record["transcript_id"],
                sequence=transcript_record["sequence"],
                config=config,
                junction_counts=junction_counts,
                reference_records=reference_records,
                cross_hybridization_index=cross_hybridization_index,
            )
            rejected_rows.extend(rejected_gene)

            selected_gene = self._select_candidate_set(
                candidates=candidates_gene,
                set_size_opt=config["set_size_opt"],
                distance_between_target_probes=config["distance_between_target_probes"],
            )

            if len(selected_gene) < config["set_size_min"]:
                rejected_rows.append(
                    {
                        "Gene": gene,
                        "Transcript_ID": transcript_record["transcript_id"],
                        "Window_Start": "",
                        "Window_End": "",
                        "Reason": f"insufficient_selected_candidates<{config['set_size_min']}",
                    }
                )
                continue

            selected_candidates.extend(selected_gene)

        outputs = self._write_outputs(
            config=config,
            selected_candidates=selected_candidates,
            assignments=assignments,
            assignment_rows=assignment_rows,
            rejected_rows=rejected_rows,
        )

        logging.info("--------------END PIPELINE--------------")
        return outputs

    ############################################
    # input loading
    ############################################

    def _load_transcript_records(self, files_fasta: list[str], gene_ids: list[str] | None) -> dict:
        records = {}
        gene_filter = set(gene_ids) if gene_ids else None

        for file_fasta in files_fasta:
            for record in SeqIO.parse(file_fasta, "fasta"):
                gene = record.id.split("::")[0]
                if gene_filter and gene not in gene_filter:
                    continue
                records[gene] = {
                    "transcript_id": record.id,
                    "sequence": str(record.seq).upper(),
                }

        if gene_filter:
            missing = sorted(gene_filter - set(records.keys()))
            for gene in missing:
                warnings.warn(f"Gene {gene} was not found in target FASTA input.", UserWarning)

        logging.info("Loaded %s transcript records for target selection.", len(records))
        return records

    def _load_reference_sequences(self, files_fasta: list[str]) -> list[dict]:
        sequences = []
        for file_fasta in files_fasta:
            for record in SeqIO.parse(file_fasta, "fasta"):
                sequences.append({"transcript_id": record.id, "sequence": str(record.seq).upper()})
        logging.info("Loaded %s reference sequences for junction counting.", len(sequences))
        return sequences

    def _load_barcode_library(self, barcode_library_config: dict) -> dict:
        workbook = load_workbook(barcode_library_config["file"], data_only=True)
        allowed_sheets = set(barcode_library_config["sheets"])
        library = {}

        for sheet_name in allowed_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Configured barcode sheet {sheet_name} is missing from workbook.")

            sheet = workbook[sheet_name]
            left_map = {}
            right_map = {}
            for row in sheet.iter_rows(values_only=True):
                if not row or len(row) < 4:
                    continue
                left_name, left_seq, right_name, right_seq = row[:4]
                if left_name and left_seq:
                    left_map[str(left_name)] = str(left_seq)
                if right_name and right_seq:
                    right_map[str(right_name)] = str(right_seq)

            library[sheet_name] = {"left": left_map, "right": right_map}

        logging.info("Loaded barcode library from %s allowed sheets.", len(library))
        return library

    def _load_assignments(
        self,
        barcode_assignment_config: dict,
        barcode_library: dict,
        gene_ids: list[str] | None,
    ) -> dict:
        with open(barcode_assignment_config["file"]) as handle:
            rows = list(csv.DictReader(handle))

        assignments = {}
        used = set()
        gene_filter = set(gene_ids) if gene_ids else None

        for row in rows:
            gene = row[barcode_assignment_config["gene_column"]]
            if gene_filter and gene not in gene_filter:
                continue

            sheet = row[barcode_assignment_config["sheet_column"]]
            left_name = row[barcode_assignment_config["left_barcode_name_column"]]
            right_name = row[barcode_assignment_config["right_barcode_name_column"]]

            if sheet not in barcode_library:
                raise ValueError(f"Assignment for {gene} refers to unknown or disallowed sheet {sheet}.")
            if left_name not in barcode_library[sheet]["left"]:
                raise ValueError(f"Assignment for {gene} refers to missing left barcode {left_name} on {sheet}.")
            if right_name not in barcode_library[sheet]["right"]:
                raise ValueError(
                    f"Assignment for {gene} refers to missing right barcode {right_name} on {sheet}."
                )

            combo = (sheet, left_name, right_name)
            if barcode_assignment_config.get("require_unique_barcode_combination", True) and combo in used:
                raise ValueError(f"Duplicate barcode combination assigned more than once: {combo}.")
            used.add(combo)

            assignments[gene] = {
                "Sheet": sheet,
                "Left_Barcode_Name": left_name,
                "Right_Barcode_Name": right_name,
                "left_sequence": barcode_library[sheet]["left"][left_name],
                "right_sequence": barcode_library[sheet]["right"][right_name],
            }

        logging.info("Loaded %s explicit barcode assignments.", len(assignments))
        return assignments

    ############################################
    # candidate generation
    ############################################

    def _build_junction_counts(self, reference_records: list[dict], cycle_hcr_config: dict) -> dict:
        junction_length = cycle_hcr_config["junction_length"]
        counts = {}
        for record in reference_records:
            sequence = record["sequence"]
            if len(sequence) < junction_length:
                continue
            for idx in range(len(sequence) - junction_length + 1):
                junction = sequence[idx : idx + junction_length]
                counts[junction] = counts.get(junction, 0) + 1
        logging.info("Indexed %s unique junction sequences of length %s.", len(counts), junction_length)
        return counts

    def _build_cross_hybridization_index(
        self, reference_records: list[dict], oligo_length: int, seed_length: int
    ) -> dict:
        index = {}
        reference_map = {record["transcript_id"]: record["sequence"] for record in reference_records}

        for record in reference_records:
            transcript_id = record["transcript_id"]
            sequence = record["sequence"]
            if len(sequence) < oligo_length or seed_length > oligo_length:
                continue
            for start in range(len(sequence) - seed_length + 1):
                seed = sequence[start : start + seed_length]
                index.setdefault(seed, []).append((transcript_id, start))

        return {"seed_length": seed_length, "seed_to_positions": index, "reference_map": reference_map}

    def _enumerate_candidates_for_gene(
        self,
        gene: str,
        transcript_id: str,
        sequence: str,
        config: dict,
        junction_counts: dict,
        reference_records: list[dict],
        cross_hybridization_index: dict | None,
    ) -> tuple[list[CandidateWindow], list[dict]]:
        cycle_hcr = config["cycle_hcr"]
        window_length = cycle_hcr["window_length"]
        left_len = cycle_hcr["left_probe_length"]
        gap_len = cycle_hcr["inter_probe_gap_length"]
        right_len = cycle_hcr["right_probe_length"]
        junction_len = cycle_hcr["junction_length"]
        junction_left_span = (junction_len - gap_len) // 2
        junction_right_span = junction_len - gap_len - junction_left_span

        candidates = []
        rejected = []

        if len(sequence) < window_length:
            rejected.append(
                {
                    "Gene": gene,
                    "Transcript_ID": transcript_id,
                    "Window_Start": "",
                    "Window_End": "",
                    "Reason": f"sequence_shorter_than_window<{window_length}",
                }
            )
            return candidates, rejected

        for idx in range(len(sequence) - window_length + 1):
            window = sequence[idx : idx + window_length]
            left = window[:left_len]
            gap = window[left_len : left_len + gap_len]
            right = window[left_len + gap_len :]
            window_start = idx + 1
            window_end = idx + window_length

            reject_reason = self._evaluate_window(left=left, right=right, config=config)
            if reject_reason:
                rejected.append(
                    {
                        "Gene": gene,
                        "Transcript_ID": transcript_id,
                        "Window_Start": window_start,
                        "Window_End": window_end,
                        "Reason": reject_reason,
                    }
                )
                continue

            max_cross_hybridization_tm_left = None
            max_cross_hybridization_tm_right = None
            if self._cross_hybridization_tm_screen_enabled(config):
                max_cross_hybridization_tm_left = self._max_cross_hybridization_tm(
                    sequence=left,
                    transcript_id=transcript_id,
                    on_target_start=window_start - 1,
                    cross_hybridization_index=cross_hybridization_index,
                    tm_parameters=config["target_probe_Tm_parameters"],
                )
                if (
                    max_cross_hybridization_tm_left is not None
                    and max_cross_hybridization_tm_left > cycle_hcr["max_cross_hybridization_tm"]
                ):
                    rejected.append(
                        {
                            "Gene": gene,
                            "Transcript_ID": transcript_id,
                            "Window_Start": window_start,
                            "Window_End": window_end,
                            "Reason": "left_cross_hybridization_tm_above_max",
                        }
                    )
                    continue

                max_cross_hybridization_tm_right = self._max_cross_hybridization_tm(
                    sequence=right,
                    transcript_id=transcript_id,
                    on_target_start=window_start - 1 + left_len + gap_len,
                    cross_hybridization_index=cross_hybridization_index,
                    tm_parameters=config["target_probe_Tm_parameters"],
                )
                if (
                    max_cross_hybridization_tm_right is not None
                    and max_cross_hybridization_tm_right > cycle_hcr["max_cross_hybridization_tm"]
                ):
                    rejected.append(
                        {
                            "Gene": gene,
                            "Transcript_ID": transcript_id,
                            "Window_Start": window_start,
                            "Window_End": window_end,
                            "Reason": "right_cross_hybridization_tm_above_max",
                        }
                    )
                    continue

            junction = left[-junction_left_span:] + gap + right[:junction_right_span]
            if junction_counts.get(junction, 0) > cycle_hcr["junction_max_reference_matches"]:
                rejected.append(
                    {
                        "Gene": gene,
                        "Transcript_ID": transcript_id,
                        "Window_Start": window_start,
                        "Window_End": window_end,
                        "Reason": "junction_not_unique_in_reference",
                    }
                )
                continue

            gc_left = self._gc_content(left)
            gc_right = self._gc_content(right)
            tm_left = self._tm_dna(left, config["target_probe_Tm_parameters"])
            tm_right = self._tm_dna(right, config["target_probe_Tm_parameters"])
            rna_tm_left = tm_left + cycle_hcr["dna_rna_tm_estimate_offset"]
            rna_tm_right = tm_right + cycle_hcr["dna_rna_tm_estimate_offset"]
            dg_left = self._secondary_structure_dg(left, config["target_probe_T_secondary_structure"])
            dg_right = self._secondary_structure_dg(right, config["target_probe_T_secondary_structure"])
            score = (
                abs(gc_left - config["target_probe_GC_content_opt"])
                + abs(gc_right - config["target_probe_GC_content_opt"])
                + abs(tm_left - config["target_probe_Tm_opt"])
                + abs(tm_right - config["target_probe_Tm_opt"])
            )

            candidates.append(
                CandidateWindow(
                    gene=gene,
                    transcript_id=transcript_id,
                    window_start=window_start,
                    window_end=window_end,
                    target_left_45=left,
                    target_gap_2=gap,
                    target_right_45=right,
                    gc_left=gc_left,
                    gc_right=gc_right,
                    tm_left=tm_left,
                    tm_right=tm_right,
                    rna_tm_left=rna_tm_left,
                    rna_tm_right=rna_tm_right,
                    dg_left=dg_left,
                    dg_right=dg_right,
                    max_cross_hybridization_tm_left=max_cross_hybridization_tm_left,
                    max_cross_hybridization_tm_right=max_cross_hybridization_tm_right,
                    junction_sequence=junction,
                    score=score,
                )
            )

        return candidates, rejected

    def _evaluate_window(self, left: str, right: str, config: dict) -> str | None:
        cycle_hcr = config["cycle_hcr"]

        for side_name, sequence in [("left", left), ("right", right)]:
            if self._contains_homopolymer(sequence, config["target_probe_homopolymeric_base_n"]):
                return f"{side_name}_homopolymer"

            gc = self._gc_content(sequence)
            if gc < config["target_probe_GC_content_min"] or gc > config["target_probe_GC_content_max"]:
                return f"{side_name}_gc_out_of_range"

            tm_dna = self._tm_dna(sequence, config["target_probe_Tm_parameters"])
            if tm_dna + cycle_hcr["dna_rna_tm_estimate_offset"] < cycle_hcr["min_dna_rna_tm_estimate"]:
                return f"{side_name}_rna_tm_below_min"

            dg_value = self._secondary_structure_dg(sequence, config["target_probe_T_secondary_structure"])
            if (
                dg_value is not None
                and dg_value < config["target_probe_secondary_structures_threshold_deltaG"]
            ):
                return f"{side_name}_secondary_structure"

        return None

    def _cross_hybridization_tm_screen_enabled(self, config: dict) -> bool:
        return bool(
            config.get("target_probe_apply_cross_hybridization", False)
            and config["cycle_hcr"].get("enable_cross_hybridization_tm_screen", False)
        )

    def _max_cross_hybridization_tm(
        self,
        sequence: str,
        transcript_id: str,
        on_target_start: int,
        cross_hybridization_index: dict,
        tm_parameters: dict,
    ) -> float | None:
        seed_length = cross_hybridization_index["seed_length"]
        seed_to_positions = cross_hybridization_index["seed_to_positions"]
        reference_map = cross_hybridization_index["reference_map"]
        candidate_windows = set()

        for seed_offset in range(len(sequence) - seed_length + 1):
            seed = sequence[seed_offset : seed_offset + seed_length]
            for matched_transcript_id, matched_seed_start in seed_to_positions.get(seed, []):
                window_start = matched_seed_start - seed_offset
                if window_start < 0:
                    continue
                if window_start + len(sequence) > len(reference_map[matched_transcript_id]):
                    continue
                if matched_transcript_id == transcript_id and window_start == on_target_start:
                    continue
                candidate_windows.add((matched_transcript_id, window_start))

        if not candidate_windows:
            return None

        max_tm = None
        for matched_transcript_id, window_start in candidate_windows:
            off_target_window = reference_map[matched_transcript_id][
                window_start : window_start + len(sequence)
            ]
            duplex_tm = self._estimate_cross_hybridization_tm(sequence, off_target_window, tm_parameters)
            if duplex_tm is None:
                continue
            if max_tm is None or duplex_tm > max_tm:
                max_tm = duplex_tm

        return max_tm

    def _select_candidate_set(
        self, candidates: list[CandidateWindow], set_size_opt: int, distance_between_target_probes: int
    ) -> list[CandidateWindow]:
        selected = []
        min_start_distance = 92 + distance_between_target_probes

        for candidate in sorted(candidates, key=lambda candidate: (candidate.score, candidate.window_start)):
            if any(abs(candidate.window_start - current.window_start) < min_start_distance for current in selected):
                continue
            selected.append(candidate)
            if len(selected) >= set_size_opt:
                break

        return sorted(selected, key=lambda candidate: candidate.window_start)

    ############################################
    # sequence helpers
    ############################################

    def _tm_dna(self, sequence: str, tm_parameters: dict) -> float:
        parameters = dict(tm_parameters)
        parameters["nn_table"] = getattr(mt, parameters["nn_table"])
        parameters["tmm_table"] = getattr(mt, parameters["tmm_table"])
        parameters["imm_table"] = getattr(mt, parameters["imm_table"])
        parameters["de_table"] = getattr(mt, parameters["de_table"])
        return float(mt.Tm_NN(sequence, **parameters))

    def _estimate_cross_hybridization_tm(
        self, sequence: str, off_target_window: str, tm_parameters: dict
    ) -> float | None:
        exact_tm = self._tm_dna(sequence, tm_parameters)
        if sequence == off_target_window:
            return exact_tm

        matches = sum(base1 == base2 for base1, base2 in zip(sequence, off_target_window))
        if matches == 0:
            return 0.0

        return exact_tm * (matches / len(sequence))

    def _secondary_structure_dg(self, sequence: str, temperature: float) -> float | None:
        if seqfold_dg is None:
            return None
        return float(seqfold_dg(sequence, temp=temperature))

    def _gc_content(self, sequence: str) -> float:
        gc = sequence.count("G") + sequence.count("C")
        return 100 * gc / len(sequence)

    def _contains_homopolymer(self, sequence: str, rules: dict) -> bool:
        for base, threshold in rules.items():
            if base * int(threshold) in sequence:
                return True
        return False

    def _reverse_complement(self, sequence: str) -> str:
        return str(Seq(sequence).reverse_complement())

    ############################################
    # output writing
    ############################################

    def _write_outputs(
        self,
        config: dict,
        selected_candidates: list[CandidateWindow],
        assignments: dict,
        assignment_rows: list[dict],
        rejected_rows: list[dict],
    ) -> dict[str, str]:
        output_dir = Path(self.dir_output)
        output_dir.mkdir(parents=True, exist_ok=True)

        panel_rows = []
        pair_rows = []

        for candidate in selected_candidates:
            assignment = assignments[candidate.gene]
            pair_id = f"{candidate.gene}_pair_{candidate.window_start}"

            direct_left = (
                self._reverse_complement(candidate.target_left_45)
                + config["direct_primary"]["spacer"]
                + assignment["left_sequence"]
            )
            direct_right = (
                assignment["right_sequence"]
                + config["direct_primary"]["spacer"]
                + self._reverse_complement(candidate.target_right_45)
            )

            twist_config = config["twist_pcr_t7_rt"]
            left_barcode_14 = assignment["left_sequence"][-twist_config["left_barcode_subsequence"]["length"] :]
            right_barcode_14 = assignment["right_sequence"][
                : twist_config["right_barcode_subsequence"]["length"]
            ]
            twist_left = (
                twist_config["forward_primer_with_t7"]
                + candidate.target_left_45
                + twist_config["spacer"]
                + left_barcode_14
                + twist_config["reverse_primer_sequence"]
            )
            twist_right = (
                twist_config["forward_primer_with_t7"]
                + right_barcode_14
                + twist_config["spacer"]
                + candidate.target_right_45
                + twist_config["reverse_primer_sequence"]
            )

            pair_rows.append(
                {
                    "Gene": candidate.gene,
                    "Probe_Pair_ID": pair_id,
                    "Transcript_ID": candidate.transcript_id,
                    "Window_Start": candidate.window_start,
                    "Window_End": candidate.window_end,
                    "Target_Left_45": candidate.target_left_45,
                    "Target_Gap_2": candidate.target_gap_2,
                    "Target_Right_45": candidate.target_right_45,
                    "Junction_Sequence": candidate.junction_sequence,
                    "Barcode_Sheet": assignment["Sheet"],
                    "Left_Barcode_Name": assignment["Left_Barcode_Name"],
                    "Right_Barcode_Name": assignment["Right_Barcode_Name"],
                    "Left_Direct_Sequence": direct_left,
                    "Right_Direct_Sequence": direct_right,
                    "Left_Twist_PCR_T7_RT_Sequence": twist_left,
                    "Right_Twist_PCR_T7_RT_Sequence": twist_right,
                }
            )

            if "direct" in config["output_modes"]:
                panel_rows.extend(
                    [
                        self._panel_row(
                            candidate=candidate,
                            pair_id=pair_id,
                            side="L",
                            output_mode="direct",
                            binding_sequence=self._reverse_complement(candidate.target_left_45),
                            final_sequence=direct_left,
                            assignment=assignment,
                        ),
                        self._panel_row(
                            candidate=candidate,
                            pair_id=pair_id,
                            side="R",
                            output_mode="direct",
                            binding_sequence=self._reverse_complement(candidate.target_right_45),
                            final_sequence=direct_right,
                            assignment=assignment,
                        ),
                    ]
                )

            if "twist_pcr_t7_rt" in config["output_modes"]:
                panel_rows.extend(
                    [
                        self._panel_row(
                            candidate=candidate,
                            pair_id=pair_id,
                            side="L",
                            output_mode="twist_pcr_t7_rt",
                            binding_sequence=candidate.target_left_45,
                            final_sequence=twist_left,
                            assignment=assignment,
                        ),
                        self._panel_row(
                            candidate=candidate,
                            pair_id=pair_id,
                            side="R",
                            output_mode="twist_pcr_t7_rt",
                            binding_sequence=candidate.target_right_45,
                            final_sequence=twist_right,
                            assignment=assignment,
                        ),
                    ]
                )

        outputs = {}
        outputs["panel"] = str(output_dir / "cycle_hcr_probe_panel.tsv")
        outputs["pairs"] = str(output_dir / "cycle_hcr_probe_pairs.tsv")
        outputs["assignments"] = str(output_dir / "cycle_hcr_barcode_assignments.tsv")
        outputs["rejected"] = str(output_dir / "cycle_hcr_rejected_candidates.tsv")

        panel_columns = [
            "Gene",
            "Probe_Pair_ID",
            "Side",
            "Transcript_ID",
            "Window_Start",
            "Window_End",
            "Target_Left_45",
            "Target_Right_45",
            "Binding_Sequence",
            "Barcode_Sheet",
            "Left_Barcode_Name",
            "Right_Barcode_Name",
            "Left_Barcode_Sequence",
            "Right_Barcode_Sequence",
            "Output_Mode",
            "Final_Probe_Sequence",
            "GC_Left",
            "GC_Right",
            "Tm_Left",
            "Tm_Right",
            "RNA_Tm_Left_Estimate",
            "RNA_Tm_Right_Estimate",
            "Max_Cross_Hybridization_Tm_Left",
            "Max_Cross_Hybridization_Tm_Right",
            "Junction_Sequence",
            "Candidate_Score",
        ]
        pair_columns = [
            "Gene",
            "Probe_Pair_ID",
            "Transcript_ID",
            "Window_Start",
            "Window_End",
            "Target_Left_45",
            "Target_Gap_2",
            "Target_Right_45",
            "Junction_Sequence",
            "Barcode_Sheet",
            "Left_Barcode_Name",
            "Right_Barcode_Name",
            "Left_Direct_Sequence",
            "Right_Direct_Sequence",
            "Left_Twist_PCR_T7_RT_Sequence",
            "Right_Twist_PCR_T7_RT_Sequence",
        ]
        assignment_columns = [
            "Gene",
            "Sheet",
            "Left_Barcode_Name",
            "Right_Barcode_Name",
            "Left_Barcode_Sequence",
            "Right_Barcode_Sequence",
        ]
        rejected_columns = ["Gene", "Transcript_ID", "Window_Start", "Window_End", "Reason"]

        pd.DataFrame(panel_rows, columns=panel_columns).to_csv(outputs["panel"], sep="\t", index=False)
        pd.DataFrame(pair_rows, columns=pair_columns).to_csv(outputs["pairs"], sep="\t", index=False)
        pd.DataFrame(assignment_rows, columns=assignment_columns).to_csv(
            outputs["assignments"], sep="\t", index=False
        )
        pd.DataFrame(rejected_rows, columns=rejected_columns).to_csv(
            outputs["rejected"], sep="\t", index=False
        )

        return outputs

    def _panel_row(
        self,
        candidate: CandidateWindow,
        pair_id: str,
        side: str,
        output_mode: str,
        binding_sequence: str,
        final_sequence: str,
        assignment: dict,
    ) -> dict:
        return {
            "Gene": candidate.gene,
            "Probe_Pair_ID": pair_id,
            "Side": side,
            "Transcript_ID": candidate.transcript_id,
            "Window_Start": candidate.window_start,
            "Window_End": candidate.window_end,
            "Target_Left_45": candidate.target_left_45,
            "Target_Right_45": candidate.target_right_45,
            "Binding_Sequence": binding_sequence,
            "Barcode_Sheet": assignment["Sheet"],
            "Left_Barcode_Name": assignment["Left_Barcode_Name"],
            "Right_Barcode_Name": assignment["Right_Barcode_Name"],
            "Left_Barcode_Sequence": assignment["left_sequence"],
            "Right_Barcode_Sequence": assignment["right_sequence"],
            "Output_Mode": output_mode,
            "Final_Probe_Sequence": final_sequence,
            "GC_Left": candidate.gc_left,
            "GC_Right": candidate.gc_right,
            "Tm_Left": candidate.tm_left,
            "Tm_Right": candidate.tm_right,
            "RNA_Tm_Left_Estimate": candidate.rna_tm_left,
            "RNA_Tm_Right_Estimate": candidate.rna_tm_right,
            "Max_Cross_Hybridization_Tm_Left": candidate.max_cross_hybridization_tm_left,
            "Max_Cross_Hybridization_Tm_Right": candidate.max_cross_hybridization_tm_right,
            "Junction_Sequence": candidate.junction_sequence,
            "Candidate_Score": candidate.score,
        }


############################################
# CLI
############################################


def base_parser():
    parser = ArgumentParser(
        prog="CycleHCR Probe Designer",
        usage="cycle_hcr_probe_designer [options]",
        description=__doc__,
        formatter_class=RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "-c",
        "--config",
        help="Path to the config file in yaml format, str",
        default=None,
        type=str,
        metavar="",
    )
    args = parser.parse_args()
    return vars(args)


def _read_gene_ids(file_regions: str | None) -> list[str] | None:
    if file_regions is None:
        warnings.warn(
            "No gene list file was provided. All regions from the target FASTA input will be considered.",
            UserWarning,
        )
        return None

    with open(file_regions) as handle:
        return sorted(set(line.rstrip() for line in handle if line.strip()))


def main():
    print("--------------START PIPELINE--------------")

    args = base_parser()
    with open(args["config"], "r") as handle:
        config = yaml.safe_load(handle)

    gene_ids = _read_gene_ids(config.get("file_regions"))

    pipeline = CycleHCRProbeDesigner(
        write_intermediate_steps=config["write_intermediate_steps"],
        dir_output=config["dir_output"],
        n_jobs=config["n_jobs"],
    )
    outputs = pipeline.run(config=config, gene_ids=gene_ids)

    for output_name, output_file in outputs.items():
        print(f"{output_name}: {output_file}")


if __name__ == "__main__":
    main()
